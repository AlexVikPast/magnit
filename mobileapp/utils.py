from datetime import datetime
import openpyxl
from .models import Products
from collections import Counter
from django.db.models import Min, Max, Sum


def print_t(ob):
    print(type(ob), ob)


def first_item_dict_queryset(ob):
    for _ in ob:
        for k, v in _.items():
            return v


def first_item_dict(dc):

    for k, v in dc.items():
        return v


def logic(date):

    n = 2
    dc = {}
    products = Products.objects.all()
    for product in products:
        if Products.objects.filter(title=product.title).count() >= n:
            # print('++++++++++++', product.title)
            if not product.title in dc:

                date_min = Products.objects.filter(title=product.title).aggregate(Min('purchase_date'))
                date_max = Products.objects.filter(title=product.title).aggregate(Max('purchase_date'))

                date_min_val = first_item_dict(date_min)
                date_max_val = first_item_dict(date_max)

                sum_count = Products.objects.filter(title=product.title).aggregate(Sum('count'))
                sum_count_val = first_item_dict(sum_count)

                count_last = Products.objects.values('count').filter(title=product.title, purchase_date=date_max_val)
                count_last_val = first_item_dict_queryset(count_last)

                # (date_max_val - date_min_val).days

                # sum_count_val/((date_max_val - date_min_val).days) - расход в день

                if (sum_count_val/((date_max_val - date_min_val).days))*((date-date_max_val).days) > count_last_val:
                    dc[product.title] = product.title

    return dc


def write_db(excel_data):
    import sys
    import os

    project_dir = r'C:\Users\Alex\PycharmProjects\mobile\mobile\mobileapp'
    sys.path.append(project_dir)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'settings'

    import django
    django.setup()

    from mobileapp.models import Products

    for i in excel_data:
        product = Products()

        product.title = i[0]
        product.count = i[1]
        dt = datetime.strptime(i[2], '%Y-%m-%d %H:%M:%S').date()
        product.purchase_date = dt

        product.save()


def open_data(excel_file):
    # you may put validations here to check extension or file size
    wb = openpyxl.load_workbook(excel_file)

    # getting all sheets
    sheets = wb.sheetnames
    # print(sheets)

    # getting a particular sheet
    worksheet = wb["Лист1"]
    # print(worksheet)

    # getting active sheet
    active_sheet = wb.active
    # print(active_sheet)

    # reading a cell
    # print(worksheet["A1"].value)

    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
            # print(cell.value)
        excel_data.append(row_data)
    # print(excel_data)
    write_db(excel_data)
    # return render(request, 'mobileapp/upload_page.html', {"excel_data": excel_data})


# Определяем минимум или максимум значений в словаре, и получаем на выходе словарь
def min_max_dc(dc, min_max, key_item, return_type):

    if min_max == 'min':
        val = min(dc.values())
    elif min_max == 'max':
        val = max(dc.values())

    if key_item == 'item':
        if return_type == 'dict':
            return {k: v for k, v in dc.items() if v == val}
        elif return_type == 'list':
            return [k for k, v in dc.items() if v == val]
    elif key_item == 'key':

        dc = dict(dc)
        arr = []
        for k in dc.keys():
            arr.append(k)

        arr_return = []
        dc_return = {}
        if min_max == 'min' and return_type == 'list':
            arr_return.append(min(arr))
            return arr_return
        elif min_max == 'max' and return_type == 'list':
            arr_return.append(max(arr))
            return arr_return
        elif min_max == 'min' and return_type == 'dict':
            dc_return = {min(arr): dc[min(arr)]}
            return dc_return
        elif min_max == 'max' and return_type == 'dict':
            dc_return = {max(arr): dc[max(arr)]}
            return dc_return

        if min_max == 'min':
            val = min(dc.values())
        elif min_max == 'max':
            val = max(dc.values())

        if return_type == 'dict':
            return {k: v for k, v in dc.iteritems() if v == val}
        elif return_type == 'list':
            return [k for k, v in dc.iteritems() if v == val]


def min_date():

    if Products.objects.count() > 0:
        ar_purchase_date = []
        product = Products.objects.all()

        for _ in product:
            ar_purchase_date.append(_.purchase_date)

        dc_purchase_date = Counter(ar_purchase_date)

        _date = min_max_dc(dc_purchase_date, 'max', 'key', 'list')
        return _date[0]

    else:
        return None


def prediction_main():
    # n - количество повторений в запуке товара
    n = 4

    ar_title, ar_purchase_date = [], []
    product = Products.objects.all()

    for _ in product:
        ar_title.append(_.title)
        ar_purchase_date.append(_.purchase_date)

    dc_title = Counter(ar_title)
    dc_purchase_date = Counter(ar_purchase_date)

    min_date = min_max_dc(dc_purchase_date, 'min', 'item', 'list')
    max_date = min_max_dc(dc_purchase_date, 'max', 'item', 'list')
    min_date, max_date = min_date[0], max_date[0]
    delta_day = (max_date - min_date).days
    print(delta_day)

    # for key, item in dc_title.items():
    #     if item > n:
    #         print(key, item)

    return None
